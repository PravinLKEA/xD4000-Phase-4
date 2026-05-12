
import csv, os, sys, time
from dataclasses import dataclass
from typing import Optional
from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import QApplication,QMainWindow,QWidget,QVBoxLayout,QHBoxLayout,QGridLayout,QLabel,QLineEdit,QPushButton,QTableWidget,QTableWidgetItem,QMessageBox,QSpinBox,QTextEdit,QGroupBox,QCheckBox,QFileDialog,QTabWidget,QComboBox
try:
    import pyqtgraph as pg
except Exception:
    pg=None
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except Exception:
    Workbook=load_workbook=None
try:
    from pymodbus.client import ModbusTcpClient
except Exception:
    ModbusTcpClient=None
BRAND_BLUE='#008CD7'; BRAND_BLUE_DARK='#005C8E'; BRAND_BLUE_DEEP='#004265'; BG='#EFEFEF'; CARD='#FFFFFF'; TEXT='#3C3C3C'; TEXT2='#7E7E7E'
DATA_COLORS=['#0766F6','#DC272D','#00943D','#FFDD49','#8886FB','#F2AC59','#585B5B']

def resource_path(p): return os.path.join(getattr(sys,'_MEIPASS',os.path.abspath(os.path.join(os.path.dirname(__file__),'..'))),p)
def sf(v,d=0.0):
    try: return d if v is None or str(v).strip()=='' else float(v)
    except Exception: return d
def si(v,d=0):
    try: return int(float(v))
    except Exception: return d
@dataclass
class Parameter:
    model:str; reference:str; code:str; name:str; address:int; datatype:str; scale:float; default:float; min:float; max:float; unit:str; access:str; monitor:bool; write_protect:bool=False; scope:bool=False; group:str=''; subcategory:str=''; notes:str=''; value:Optional[float]=None; online_value:Optional[float]=None; user_modified:bool=False
    @property
    def effective_value(self): return self.value if self.value is not None else self.default
class ParameterDB:
    def __init__(self): self.params=[]
    def load_csv(self,path):
        self.params=[]
        with open(path,newline='',encoding='utf-8-sig') as f:
            for r in csv.DictReader(f):
                self.params.append(Parameter((r.get('model') or 'XD4000').strip(),(r.get('reference') or 'ALL').strip(),(r.get('code') or '').strip(),(r.get('name') or '').strip(),si(r.get('address')),(r.get('datatype') or 'uint16').strip().lower(),sf(r.get('scale'),1),sf(r.get('default')),sf(r.get('min'),-32768),sf(r.get('max'),65535),(r.get('unit') or '').strip(),(r.get('access') or 'RO').strip().upper(),str(r.get('monitor') or 'FALSE').upper() in ('TRUE','1','YES','Y'),str(r.get('write_protect') or 'FALSE').upper() in ('TRUE','1','YES','Y'),str(r.get('scope') or 'FALSE').upper() in ('TRUE','1','YES','Y'),(r.get('group') or '').strip(),(r.get('subcategory') or '').strip(),(r.get('notes') or '').strip()))
    def filtered(self,search='',monitor_only=False):
        s=(search or '').lower().strip(); out=[]
        for p in self.params:
            if monitor_only and not p.monitor: continue
            if s and not (s in p.code.lower() or s in p.name.lower() or s in str(p.address) or s in p.group.lower()): continue
            out.append(p)
        return out
    def by_code(self,code):
        for p in self.params:
            if p.code.upper()==code.upper(): return p
        return None
    def scope_params(self): return [p for p in self.params if p.scope]
class ModbusGateway:
    def __init__(self): self.client=None; self.unit_id=1; self.address_offset=0
    def connect_tcp(self,host,port,unit_id,zero_based=False):
        if ModbusTcpClient is None: raise RuntimeError('pymodbus is not installed')
        self.unit_id=unit_id; self.address_offset=-1 if zero_based else 0; self.client=ModbusTcpClient(host=host,port=port,timeout=3)
        if not self.client.connect(): raise RuntimeError('Could not connect to Modbus TCP device')
    def close(self):
        if self.client: self.client.close()
        self.client=None
    def is_connected(self): return self.client is not None
    def _addr(self,a):
        a=a+self.address_offset
        if a<0: raise RuntimeError(f'Invalid address after offset: {a}')
        return a
    def _kwargs(self): return [{'slave':self.unit_id},{'unit':self.unit_id},{'device_id':self.unit_id},{}]
    def read_registers(self,address,count=1):
        address=self._addr(address); last=None
        for kw in self._kwargs():
            try:
                rr=self.client.read_holding_registers(address=address,count=count,**kw)
                if rr.isError(): raise RuntimeError(str(rr))
                return rr.registers
            except TypeError as e: last=e; continue
        raise RuntimeError(f'read_holding_registers API failed: {last}')
    def write_register(self,address,value):
        address=self._addr(address); last=None
        for kw in self._kwargs():
            try:
                wr=self.client.write_register(address=address,value=value,**kw)
                if wr.isError(): raise RuntimeError(str(wr))
                return
            except TypeError as e: last=e; continue
            except Exception as e: last=e; break
        for kw in self._kwargs():
            try:
                wr=self.client.write_registers(address=address,values=[value],**kw)
                if wr.isError(): raise RuntimeError(str(wr))
                return
            except TypeError as e: last=e; continue
            except Exception as e: last=e; break
        raise RuntimeError(f'Write failed using FC06 and FC16: {last}')
    def read_param(self,p):
        regs=self.read_registers(p.address,2 if p.datatype in ('uint32','int32') else 1)
        if p.datatype=='int16': raw=regs[0] if regs[0]<32768 else regs[0]-65536
        elif p.datatype=='uint32': raw=(regs[0]<<16)+regs[1]
        elif p.datatype=='int32':
            raw=(regs[0]<<16)+regs[1]
            if raw>=2147483648: raw-=4294967296
        else: raw=regs[0]
        return raw*p.scale
    def write_param(self,p,val):
        if p.access!='RW': raise RuntimeError(f'{p.code} is read-only')
        raw=int(round(val/(p.scale if p.scale else 1)))
        if p.datatype=='int16' and raw<0: raw=65536+raw
        self.write_register(p.address,raw & 0xFFFF)
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__(); self.setWindowTitle('LK XD4000 Phase-5A/5B Full Parameter Manager + Command Dry-Run'); self.resize(1540,920)
        self.db=ParameterDB(); self.gateway=ModbusGateway(); self.params=[]; self.last_eligibility=False
        self.keepalive_timer=QTimer(self); self.keepalive_timer.timeout.connect(self.keepalive_tick)
        self.scope_timer=QTimer(self); self.scope_timer.timeout.connect(self.scope_tick)
        self.scope_start=None; self.scope_data={}; self.scope_checks={}
        self.build_ui(); self.apply_lk_theme(); self.load_db()
    def apply_lk_theme(self):
        self.setStyleSheet(f"""QMainWindow {{background:{BG}; font-family:Roboto, Segoe UI, sans-serif; color:{TEXT};}} QGroupBox {{background:{CARD}; border:1px solid #C1C1C1; border-radius:10px; margin-top:12px; padding:10px; font-weight:bold; color:{TEXT};}} QLabel {{color:{TEXT}; font-size:12px;}} QLineEdit,QSpinBox,QComboBox {{background:#FFF; border:1px solid #C1C1C1; border-radius:6px; padding:5px; min-height:24px;}} QPushButton {{background:{BRAND_BLUE}; color:white; border:none; border-radius:8px; min-height:35px; padding:6px 12px; font-weight:bold;}} QPushButton:hover {{background:{BRAND_BLUE_DARK};}} QPushButton:disabled {{background:#C1C1C1; color:#7E7E7E;}} QTabWidget::pane {{border:1px solid #C1C1C1; background:{CARD}; border-radius:8px;}} QTabBar::tab {{background:#EFEFEF; color:{TEXT2}; padding:8px 18px; min-height:24px;}} QTabBar::tab:selected {{background:{CARD}; color:{BRAND_BLUE_DARK}; border-bottom:3px solid {BRAND_BLUE}; font-weight:bold;}} QTableWidget {{background:{CARD}; gridline-color:#EFEFEF; selection-background-color:#E6F6FC;}} QHeaderView::section {{background:#EFEFEF; color:{TEXT}; padding:6px; border:0px; font-weight:bold;}} QTextEdit {{background:#FFF; border:1px solid #C1C1C1; border-radius:8px; padding:6px;}}""")
    def log(self,m): self.logbox.append(f'[{time.strftime("%H:%M:%S")}] {m}')
    def load_db(self):
        self.db.load_csv(resource_path(os.path.join('data','xd4000_phase5_parameters.csv'))); self.refresh_params(); self.build_scope_signal_list(); self.log(f'Loaded XD4000 Phase-5 database: {len(self.db.params)} parameters')
    def build_ui(self):
        c=QWidget(); self.setCentralWidget(c); root=QVBoxLayout(c)
        appbar=QGroupBox('Lauritz Knudsen  |  XD4000 Phase-5A/5B Parameter Manager + Command Dry-Run'); appbar_layout=QHBoxLayout(appbar)
        self.connection_status=QLabel('Offline'); self.connection_status.setStyleSheet(f'color:{BRAND_BLUE_DEEP}; font-weight:bold;')
        appbar_layout.addWidget(QLabel('Drive Manager')); appbar_layout.addStretch(); appbar_layout.addWidget(self.connection_status); root.addWidget(appbar)
        box=QGroupBox('XD4000 / ATV930 - Modbus TCP'); g=QGridLayout(box)
        self.host=QLineEdit('192.168.1.10'); self.port=QSpinBox(); self.port.setRange(1,65535); self.port.setValue(502); self.unit=QSpinBox(); self.unit.setRange(1,255); self.unit.setValue(1)
        self.zero=QCheckBox('Use zero-based address (-1)'); self.search=QLineEdit(); self.search.setPlaceholderText('Search code/name/address/group'); self.search.textChanged.connect(self.refresh_params); self.mononly=QCheckBox('Monitor only'); self.mononly.stateChanged.connect(self.refresh_params)
        for i,(lab,w) in enumerate([('Drive IP',self.host),('TCP Port',self.port),('Unit ID',self.unit),('Address option',self.zero),('Search',self.search),('Filter',self.mononly)]): g.addWidget(QLabel(lab),0,i); g.addWidget(w,1,i)
        for i,(txt,fn) in enumerate([('Connect',self.connect_drive),('Disconnect',self.disconnect_drive),('Upload visible',self.upload_visible),('Download selected row',self.download_selected),('Export project Excel',self.export_project_excel),('Import project Excel',self.import_project_excel)]):
            b=QPushButton(txt); b.clicked.connect(fn); g.addWidget(b,2,i)
        root.addWidget(box); self.tabs=QTabWidget(); root.addWidget(self.tabs,1); self.table=QTableWidget(); self.tabs.addTab(self.table,'Parameters')
        self.build_ethernet_tab(); self.build_command_tab(); self.build_scope_tab(); self.logbox=QTextEdit(); self.logbox.setReadOnly(True); self.tabs.addTab(self.logbox,'Event Log')
    def build_ethernet_tab(self):
        diag=QWidget(); dl=QVBoxLayout(diag); self.expert=QCheckBox('Expert test mode: bench setup confirmed, output terminals safe'); dl.addWidget(self.expert)
        self.keepalive=QCheckBox('Maintain Ethernet keep-alive polling every 1 second'); self.keepalive.stateChanged.connect(self.toggle_keepalive); dl.addWidget(self.keepalive)
        row=QHBoxLayout()
        for txt,fn in [('Diagnose Ethernet Supervision',self.diagnose_ethernet),('Diagnose CRC/CCC Channels',self.diagnose_channels),('Command/Reference Config',self.diagnose_config),('Set LFR to 0.0 Hz',self.set_lfr_zero),('Command Test Checklist',self.prepare_checklist)]:
            b=QPushButton(txt); b.clicked.connect(fn); row.addWidget(b)
        dl.addLayout(row); self.diagbox=QTextEdit(); self.diagbox.setReadOnly(True); dl.addWidget(self.diagbox); self.tabs.addTab(diag,'Ethernet Supervision')
    def build_command_tab(self):
        tab=QWidget(); root=QVBoxLayout(tab); panel=QGroupBox('Authorized Engineer Command Console - Dry Run Only'); gl=QGridLayout(panel)
        self.cmd_status=QTextEdit(); self.cmd_status.setReadOnly(True)
        commands=[('Read Drive State',self.cmd_read_state),('Check Command Eligibility',self.cmd_check_eligibility),('Dry Run START',lambda:self.command_dry_run('START_FORWARD')),('Dry Run STOP',lambda:self.command_dry_run('STOP')),('Dry Run FAULT RESET',lambda:self.command_dry_run('FAULT_RESET'))]
        for i,(txt,fn) in enumerate(commands):
            b=QPushButton(txt); b.clicked.connect(fn); gl.addWidget(b,0,i)
        root.addWidget(panel); root.addWidget(QLabel('Dry-run only: no CMD@8501 command write is performed by this build.'))
        root.addWidget(self.cmd_status,1); self.tabs.addTab(tab,'Command Dry-Run')
    def build_scope_tab(self):
        tab=QWidget(); root=QVBoxLayout(tab); controls=QGroupBox('Oscilloscope / Trend Controls'); cl=QGridLayout(controls)
        self.scope_interval=QComboBox(); self.scope_interval.addItems(['250','500','1000']); self.scope_interval.setCurrentText('500')
        self.scope_window=QComboBox(); self.scope_window.addItems(['30','60','120']); self.scope_window.setCurrentText('60')
        widgets=[self.scope_interval,self.scope_window]
        for label,fn in [('Start capture',self.start_scope),('Stop capture',self.stop_scope),('Clear trace',self.clear_scope),('Export trend CSV',self.export_scope_csv)]:
            b=QPushButton(label); b.clicked.connect(fn); widgets.append(b)
        for i,(lab,w) in enumerate([('Sample interval ms',widgets[0]),('Rolling window s',widgets[1]),('',widgets[2]),('',widgets[3]),('',widgets[4]),('',widgets[5])]): cl.addWidget(QLabel(lab),0,i); cl.addWidget(w,1,i)
        root.addWidget(controls); signals=QGroupBox('Signals'); self.signal_layout=QHBoxLayout(signals); root.addWidget(signals)
        if pg:
            self.plot=pg.PlotWidget(title='XD4000 Oscilloscope / Trend'); self.plot.setBackground('w'); self.plot.showGrid(x=True,y=True); self.plot.addLegend(); self.plot.setLabel('bottom','Time',units='s'); self.plot.setLabel('left','Engineering value')
        else:
            self.plot=QTextEdit('pyqtgraph is not installed.'); self.plot.setReadOnly(True)
        root.addWidget(self.plot,1); self.tabs.addTab(tab,'Oscilloscope / Trend')
    def build_scope_signal_list(self):
        if not hasattr(self,'signal_layout'): return
        for cb in self.scope_checks.values(): self.signal_layout.removeWidget(cb); cb.deleteLater()
        self.scope_checks={}
        for p in self.db.scope_params():
            cb=QCheckBox(f'{p.code} ({p.unit})'); cb.setChecked(p.code in ('RFR','FRH','LFR')); self.signal_layout.addWidget(cb); self.scope_checks[p.code]=cb
        self.signal_layout.addStretch()
    def refresh_params(self): self.params=self.db.filtered(self.search.text() if hasattr(self,'search') else '', self.mononly.isChecked() if hasattr(self,'mononly') else False); self.populate()
    def fmt(self,v):
        if v=='': return ''
        try: return f'{float(v):.3f}'.rstrip('0').rstrip('.')
        except Exception: return str(v)
    def populate(self):
        heads=['Code','Group','Name','Address','Type','Scale','Default','Offline Value','Online Value','Unit','Access','Protect','Scope','Notes']; self.table.blockSignals(True); self.table.setColumnCount(len(heads)); self.table.setHorizontalHeaderLabels(heads); self.table.setRowCount(len(self.params))
        for r,p in enumerate(self.params):
            vals=[p.code,p.group,p.name,p.address,p.datatype,p.scale,p.default,p.effective_value,'' if p.online_value is None else self.fmt(p.online_value),p.unit,p.access,'Yes' if p.write_protect else 'No','Yes' if p.scope else 'No',p.notes]
            for col,v in enumerate(vals):
                it=QTableWidgetItem(str(v)); it.setFlags((it.flags()|Qt.ItemIsEditable) if col==7 and p.access=='RW' and not p.write_protect else (it.flags()&~Qt.ItemIsEditable)); self.table.setItem(r,col,it)
        self.table.blockSignals(False)
        try: self.table.itemChanged.disconnect()
        except Exception: pass
        self.table.itemChanged.connect(self.on_edit); self.table.resizeColumnsToContents()
    def on_edit(self,item):
        if item.column()!=7 or item.row()>=len(self.params): return
        p=self.params[item.row()]
        try:
            val=float(item.text())
            if not(p.min<=val<=p.max): raise ValueError(f'Allowed range: {p.min} to {p.max}')
            p.value=val; p.user_modified=True; self.log(f'Offline value changed: {p.code} = {val}')
        except Exception as e: QMessageBox.warning(self,'Invalid value',str(e))
    def connect_drive(self):
        try: self.gateway.connect_tcp(self.host.text().strip(),self.port.value(),self.unit.value(),self.zero.isChecked()); self.connection_status.setText('Online'); self.log(f'Connected successfully to {self.host.text().strip()}:{self.port.value()}, Unit ID={self.unit.value()}, zero_based={self.zero.isChecked()}')
        except Exception as e: self.log(f'Connection failed: {e}'); QMessageBox.critical(self,'Connection failed',str(e))
    def disconnect_drive(self): self.keepalive.setChecked(False); self.stop_scope(); self.gateway.close(); self.connection_status.setText('Offline'); self.log('Disconnected')
    def upload_one(self,p,update_offline=True):
        p.online_value=self.gateway.read_param(p)
        if update_offline and not p.user_modified: p.value=p.online_value
        self.log(f'Upload OK {p.code}@{p.address} = {self.fmt(p.online_value)} {p.unit}'); return p.online_value
    def upload_visible(self):
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return
        ok=fail=0
        for p in self.params:
            try: self.upload_one(p,True); ok+=1
            except Exception as e: self.log(f'Upload failed {p.code}@{p.address}: {e}'); fail+=1
        self.populate(); self.log(f'Upload complete. OK={ok}, Failed={fail}')
    def write_rb(self,p):
        if p.write_protect and not self.expert.isChecked(): raise RuntimeError(f'{p.code} is write-protected. Expert mode required.')
        if p.code.upper()=='CMD': raise RuntimeError('CMD raw command writes are blocked in Phase-5A dry-run build.')
        ka=self.keepalive_timer.isActive(); sc=self.scope_timer.isActive()
        if ka: self.keepalive_timer.stop(); self.log('Keep-alive temporarily paused for parameter write')
        if sc: self.scope_timer.stop(); self.log('Oscilloscope capture temporarily paused for parameter write')
        try:
            last=None
            for attempt in range(1,4):
                try: self.gateway.write_param(p,p.effective_value); self.log(f'Download OK {p.code}@{p.address} = {p.effective_value} {p.unit} on attempt {attempt}'); break
                except Exception as e: last=e; self.log(f'Download retry {attempt} failed {p.code}@{p.address}: {e}'); time.sleep(0.75*attempt)
            else: raise last
            rb=self.gateway.read_param(p); p.online_value=rb; p.value=rb; p.user_modified=False; self.log(f'Readback OK {p.code}@{p.address} = {self.fmt(rb)} {p.unit}')
        finally:
            if ka and self.keepalive.isChecked(): self.keepalive_timer.start(1000); self.log('Keep-alive resumed after parameter write')
            if sc: self.scope_timer.start(int(self.scope_interval.currentText())); self.log('Oscilloscope capture resumed after parameter write')
    def selected_param(self):
        r=self.table.currentRow(); return self.params[r] if 0<=r<len(self.params) else None
    def download_selected(self):
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return
        p=self.selected_param()
        if not p: QMessageBox.warning(self,'No row selected','Select one parameter row first'); return
        if p.access!='RW': QMessageBox.warning(self,'Read-only',f'{p.code} is read-only'); return
        if QMessageBox.question(self,'Confirm selected write',f'Write {p.code}@{p.address} = {p.effective_value} {p.unit}?')!=QMessageBox.Yes: self.log('Selected write cancelled'); return
        try: self.write_rb(p); self.populate()
        except Exception as e: self.log(f'Selected write failed {p.code}@{p.address}: {e}')
    def bit_text(self,val):
        labels={0:'Terminal/local',1:'Local keypad',2:'Remote keypad',3:'Serial Modbus / RTU',6:'CANopen',9:'Fieldbus/comm module',11:'Embedded Ethernet / Modbus TCP',15:'SoMove/PC tool'}
        active=[f'bit {b}: {lab}' for b,lab in labels.items() if int(val)&(1<<b)]
        return ', '.join(active) if active else 'no known channel bit active'
    def read_codes(self,codes,update_offline=False):
        lines=[]; vals={}
        for code in codes:
            p=self.db.by_code(code)
            if not p: lines.append(f'{code}: not in database'); continue
            try:
                v=self.gateway.read_param(p); p.online_value=v; vals[code]=v
                if update_offline and not p.user_modified: p.value=v
                lines.append(f'{code}@{p.address} = {self.fmt(v)} {p.unit}')
            except Exception as e: lines.append(f'{code}@{p.address} failed: {e}')
        return lines,vals
    def add_diag(self,title,lines): self.diagbox.append(f'[{time.strftime("%H:%M:%S")}] {title}\n'+'\n'.join(lines)+'\n')
    def diagnose_channels(self):
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return
        lines,vals=self.read_codes(['ETA','HMIS','CRC','CCC','CNFS','LFT','COM1','RFR','FRH','LFR'])
        if 'CRC' in vals: lines.append(f'CRC active reference decode: {self.bit_text(vals["CRC"])}')
        if 'CCC' in vals: lines.append(f'CCC active command decode: {self.bit_text(vals["CCC"])}')
        self.add_diag('CHANNEL DIAGNOSIS',lines); self.log('Channel diagnosis completed'); self.refresh_params()
    def diagnose_ethernet(self):
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return
        lines,vals=self.read_codes(['ETHL','ETHF','TTOB','COM1','CRC','CCC','LFT','HMIS','ETA','RFR','FRH','LFR'])
        self.add_diag('ETHERNET SUPERVISION DIAGNOSIS',lines); self.log('Ethernet supervision diagnosis completed'); self.refresh_params()
    def diagnose_config(self):
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return
        lines,vals=self.read_codes(['CHCF','FR1','CD1','CD2','CRC','CCC','CNFS']); lines.append('Read-only configuration diagnosis.')
        self.add_diag('COMMAND / REFERENCE CONFIGURATION',lines); self.log('Command/reference configuration diagnosis completed'); self.refresh_params()
    def prepare_checklist(self):
        msg='Phase-5A is dry-run only. No CMD@8501 write is performed. Verify CCC/CRC/keep-alive/LFR/RFR before future command execution.\n'
        self.diagbox.append(f'[{time.strftime("%H:%M:%S")}]\n{msg}')
    def set_lfr_zero(self):
        p=self.db.by_code('LFR')
        if p: p.value=0.0; p.user_modified=True; self.search.setText('LFR'); self.refresh_params(); self.log('Prepared LFR offline value = 0.0 Hz. Use Download selected row to write if safe.')
    def toggle_keepalive(self):
        if self.keepalive.isChecked():
            if not self.gateway.is_connected(): self.keepalive.setChecked(False); QMessageBox.warning(self,'Not connected','Connect first'); return
            self.keepalive_timer.start(1000); self.log('Ethernet keep-alive polling started at 1 s interval')
        else:
            if self.keepalive_timer.isActive(): self.keepalive_timer.stop(); self.log('Ethernet keep-alive polling stopped')
    def keepalive_tick(self):
        if not self.gateway.is_connected(): self.keepalive.setChecked(False); return
        self.read_codes(['ETA','RFR','FRH','CRC','CCC','COM1','LFR'],update_offline=False); self.log('Keep-alive poll OK')
    def cmd_read_state(self):
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return
        lines,vals=self.read_codes(['ETA','HMIS','CRC','CCC','RFR','FRH','LFR','CHCF','FR1','CD1','CD2'],update_offline=False)
        self.cmd_status.append(f'[{time.strftime("%H:%M:%S")}] DRIVE STATE\n'+'\n'.join(lines)+'\n')
    def cmd_check_eligibility(self):
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return False
        lines,vals=self.read_codes(['ETA','HMIS','CRC','CCC','RFR','FRH','LFR'],update_offline=False)
        checks=[('Expert mode checked',self.expert.isChecked()),('Keep-alive active',self.keepalive_timer.isActive()),('Reference Ethernet/Modbus TCP','CRC' in vals and (int(vals['CRC'])&(1<<11))),('Command Ethernet/Modbus TCP','CCC' in vals and (int(vals['CCC'])&(1<<11))),('RFR near zero','RFR' in vals and abs(float(vals['RFR']))<=0.2),('LFR <= 10 Hz','LFR' in vals and abs(float(vals['LFR']))<=10.0)]
        self.last_eligibility=all(v for _,v in checks)
        status=['PASS: '+n if v else 'BLOCK: '+n for n,v in checks]
        status.append('DRY-RUN ELIGIBILITY: '+('PASS' if self.last_eligibility else 'BLOCKED'))
        self.cmd_status.append(f'[{time.strftime("%H:%M:%S")}] COMMAND DRY-RUN ELIGIBILITY\n'+'\n'.join(lines+status)+'\n')
        return self.last_eligibility
    def command_dry_run(self,name):
        ok=self.cmd_check_eligibility(); result='would be allowed if active command module is later enabled' if ok else 'blocked by interlocks'
        self.cmd_status.append(f'[{time.strftime("%H:%M:%S")}] DRY RUN {name}: {result}. No CMD write executed.\n')
        self.log(f'Dry run {name}: {result}')
    def start_scope(self):
        if pg is None: QMessageBox.warning(self,'Missing dependency','pyqtgraph is not installed.'); return
        if not self.gateway.is_connected(): QMessageBox.warning(self,'Not connected','Connect first'); return
        self.scope_start=time.time(); self.scope_data={code:[] for code,cb in self.scope_checks.items() if cb.isChecked()}; self.scope_timer.start(int(self.scope_interval.currentText())); self.log('Oscilloscope capture started')
    def stop_scope(self):
        if self.scope_timer.isActive(): self.scope_timer.stop(); self.log('Oscilloscope capture stopped')
    def clear_scope(self):
        self.scope_data={}
        if pg and hasattr(self,'plot'): self.plot.clear(); self.plot.addLegend()
        self.log('Oscilloscope traces cleared')
    def scope_tick(self):
        if not self.gateway.is_connected(): self.stop_scope(); return
        t=time.time()-(self.scope_start or time.time()); window=float(self.scope_window.currentText())
        for idx,(code,cb) in enumerate(self.scope_checks.items()):
            if not cb.isChecked(): continue
            p=self.db.by_code(code)
            if not p: continue
            try:
                v=self.gateway.read_param(p); p.online_value=v; self.scope_data.setdefault(code,[]).append((t,v)); self.scope_data[code]=[(x,y) for x,y in self.scope_data[code] if t-x<=window]
            except Exception as e: self.log(f'Oscilloscope read failed {code}@{p.address}: {e}')
        self.update_scope_plot()
    def update_scope_plot(self):
        if pg is None: return
        self.plot.clear(); self.plot.addLegend()
        for idx,(code,pts) in enumerate(self.scope_data.items()):
            if not pts: continue
            xs,ys=zip(*pts); color=DATA_COLORS[idx%len(DATA_COLORS)]; self.plot.plot(list(xs),list(ys),pen=pg.mkPen(color=color,width=2),name=code)
    def export_scope_csv(self):
        if not self.scope_data: QMessageBox.information(self,'No data','No oscilloscope data to export.'); return
        path,_=QFileDialog.getSaveFileName(self,'Export oscilloscope CSV','xd4000_scope.csv','CSV Files (*.csv)')
        if not path: return
        with open(path,'w',newline='',encoding='utf-8') as f:
            w=csv.writer(f); w.writerow(['signal','time_s','value'])
            for code,pts in self.scope_data.items():
                for t,v in pts: w.writerow([code,f'{t:.3f}',v])
        self.log(f'Oscilloscope data exported: {path}')
    def export_project_excel(self):
        if Workbook is None: QMessageBox.warning(self,'Missing dependency','openpyxl not installed'); return
        path,_=QFileDialog.getSaveFileName(self,'Export project Excel','XD4000_Project_Parameters.xlsx','Excel Files (*.xlsx)')
        if not path: return
        wb=Workbook(); ws=wb.active; ws.title='Parameters'
        headers=['code','name','address','datatype','scale','default','offline_value','online_value','unit','access','write_protect','group','subcategory','notes']
        ws.append(headers)
        for p in self.db.params: ws.append([p.code,p.name,p.address,p.datatype,p.scale,p.default,p.effective_value,'' if p.online_value is None else p.online_value,p.unit,p.access,p.write_protect,p.group,p.subcategory,p.notes])
        info=wb.create_sheet('ProjectInfo'); info.append(['Project','XD4000 Phase-5 Parameter Project']); info.append(['Command mode','Dry-run only'])
        for cell in ws[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill('solid',fgColor='008CD7')
        wb.save(path); self.log(f'Project Excel exported: {path}')
    def import_project_excel(self):
        if load_workbook is None: QMessageBox.warning(self,'Missing dependency','openpyxl not installed'); return
        path,_=QFileDialog.getOpenFileName(self,'Import project Excel','','Excel Files (*.xlsx)')
        if not path: return
        wb=load_workbook(path,data_only=True)
        if 'Parameters' not in wb.sheetnames: QMessageBox.warning(self,'Invalid file','Parameters sheet not found'); return
        ws=wb['Parameters']; headers=[c.value for c in ws[1]]; idx={h:i for i,h in enumerate(headers)}; count=0
        for row in ws.iter_rows(min_row=2,values_only=True):
            code=row[idx.get('code',0)] if 'code' in idx else None
            if not code: continue
            p=self.db.by_code(str(code))
            if p and 'offline_value' in idx and row[idx['offline_value']] not in (None,''):
                try: p.value=float(row[idx['offline_value']]); p.user_modified=True; count+=1
                except Exception: pass
        self.refresh_params(); self.log(f'Project Excel imported. Offline values updated: {count}')
    def export_log(self):
        path,_=QFileDialog.getSaveFileName(self,'Export event log','xd4000_event_log.txt','Text Files (*.txt)')
        if path:
            extra='\n\n--- ETHERNET DIAGNOSTIC OUTPUT ---\n'+self.diagbox.toPlainText()+'\n\n--- COMMAND DRY-RUN OUTPUT ---\n'+self.cmd_status.toPlainText()
            open(path,'w',encoding='utf-8').write(self.logbox.toPlainText()+extra); self.log(f'Event log exported: {path}')
if __name__=='__main__':
    app=QApplication(sys.argv); w=MainWindow(); w.show(); sys.exit(app.exec())
